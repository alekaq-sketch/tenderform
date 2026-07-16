"""
Заполняет шаблоны расчёта тендера, не трогая формулы (только "входные" ячейки).

Тендер состоит из одного или нескольких ЛОТОВ. Каждый лот - это полная копия
шаблона (шапка + позиции + итоги), одна под другой в одном листе Excel:
  fill_lots_kz()       - template_kz-kz.xlsx      (закупка в Казахстане -> продажа в Казахстане)
  fill_lots_foreign()  - template_foreign.xlsx     (закупка за рубежом -> продажа в Казахстане)

Внутри лота может быть несколько позиций товара - они добавляются строками
внутри блока этого лота (как и раньше), а сам блок лота при этом растёт вниз
и сдвигает все лоты, идущие после него.

--------------------------------------------------------------------------------------------------
Карта входных ячеек внутри ОДНОГО блока лота, лист "расчет", template_kz-kz.xlsx
(номера строк ниже - для первого лота; для второго и следующих лотов весь блок
сдвинут вниз, но относительное расположение ячеек внутри блока то же самое):
    B2  - ФИО менеджера / закупщика
    B3  - Поставщик
    R2  - Номер лота
    B4  - Дата расчёта
    B5  - Дата начала лота
    B6  - Дата окончания лота
    B7  - Срок производства, дней
    J9  - Коэффициент наценки (> 1, например 1.5 = +50%)
    K9  - Курс USD
    H18 - Сумма дорожных расходов
    на каждый товар (строка 12, 13, ...):
        C - Наименование
        D - Количество
        E - Цена закупки DDP (с НДС)
        T - Доп.расходы (прочее), тнг - по умолчанию 0

--------------------------------------------------------------------------------------------------
Карта входных ячеек внутри ОДНОГО блока лота, лист "расчет", template_foreign.xlsx:
    B2   - ФИО менеджера / закупщика
    B3   - Поставщик
    AE2  - Номер лота
    B4   - Дата расчёта
    B5   - Дата начала лота
    B6   - Дата окончания лота
    B7   - Срок производства, дней
    B8   - Срок поставки, дней
    H9   - Дорога, всего по лоту, В ВАЛЮТЕ ЗАКУПКИ (делится между товарами пропорционально сумме)
    L9   - Коэфф. наценки (DAP -> Вход DAP, > 1)
    O9   - НДС на ввоз, % (например 0.16)
    AD8  - Валюта закупки (USD/RUB/EUR) - только для наглядности в файле
    AF9  - Курс валюты закупки к тенге на дату расчёта
    на каждый товар (строка 12, 13, ...):
        C  - Наименование
        D  - Ед. изм.
        E  - Количество
        F  - Цена закупки FCA, В ВАЛЮТЕ ЗАКУПКИ (не в тенге - см. примечание выше)
        P  - Накладные, В ВАЛЮТЕ ЗАКУПКИ (по умолчанию 0 - нет единого стандарта)
        AA - Цена продажи без НДС, тнг/шт (это уже цена для клиента в Казахстане)
        AG - Ставка пошлины, % (своя на каждый товар)
        AH - Кол-во машин / деклараций ГТД (своё на каждый товар: 1 машина = 1 декларация)
        AI - Доп.расходы (прочее), В ВАЛЮТЕ ЗАКУПКИ (по умолчанию 0) - входит в
             себестоимость DDP наравне с "Накладные", а не отдельной строкой в прибыли
        country/tnved/transport - своё на каждый товар (см. B19/B20/B21 ниже)

    B19 - Страна происхождения (по всем товарам, если совпадает; иначе построчная разбивка)
    B20 - ТН ВЭД (аналогично)
    B21 - Кол-во подвижных / вид транспорта (аналогично)
"""
import re
import openpyxl
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

_CELL_REF_RE = re.compile(r'(\$?)([A-Za-z]{1,3})(\$?)(\d+)')


def _shift_formula_rows(formula: str, insertion_row: int, shift: int) -> str:
    """openpyxl's ws.insert_rows() physically moves cells down but does NOT
    rewrite any formula's cell references anywhere in the workbook (this is
    a documented openpyxl limitation, unlike Excel's own 'insert row' which
    fixes every reference automatically). This reproduces that missing half:
    any reference whose row is >= insertion_row gets bumped by `shift`,
    wherever in the sheet the formula lives - including formulas that moved
    together with their row, so relationships within a shifted block stay
    internally consistent."""
    def repl(m):
        col_dollar, col, row_dollar, row_s = m.groups()
        row = int(row_s)
        if row >= insertion_row:
            row += shift
        return f"{col_dollar}{col}{row_dollar}{row}"
    return _CELL_REF_RE.sub(repl, formula)


def _shift_all_formulas_below(ws, insertion_row: int, shift: int, max_row: int, max_col: int):
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = _shift_formula_rows(cell.value, insertion_row, shift)


def compose_item_name(item: dict) -> str:
    return item.get("item_name") or item.get("name") or ""


def _compose_lot_note(items: list[dict], key: str) -> str:
    """Страна/ТНВЭД/транспорт бывают разными у разных товаров одного лота
    (как и цена). Если у всех товаров значение одинаковое (или задано только
    у одного товара в лоте) - пишем его одной строкой, как раньше. Если
    значения расходятся - пишем построчную разбивку по номеру позиции, чтобы
    не потерять данные ни одного товара."""
    values = [str(item.get(key) or "").strip() for item in items]
    if not any(values):
        return ""
    if len(set(values)) == 1:
        return values[0]
    return "; ".join(f"{i + 1}) {v or '-'}" for i, v in enumerate(values))


def _copy_row_style(ws, src_row: int, dst_row: int, min_col: int, max_col: int):
    """Copies formulas (translated for the new row) + full cell style from
    src_row to dst_row, for every column in [min_col, max_col]. Used when a
    lot has more than one item and we need extra rows that behave exactly
    like the template's first item row."""
    for col_idx in range(min_col, max_col + 1):
        col = get_column_letter(col_idx)
        src = ws[f"{col}{src_row}"]
        dst = ws[f"{col}{dst_row}"]
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = src.value
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def _copy_block(dst_ws, src_ws, src_top: int, src_bottom: int, dst_top: int,
                 min_col: int, max_col: int):
    """Copies a whole row range [src_top, src_bottom] - values, formulas
    (translated), styles, row heights and merged-cell ranges - from src_ws
    into dst_ws starting at dst_top. Used to stamp a fresh, pristine copy of
    the one-lot template at a new position for each additional lot in a
    multi-lot tender, so every lot's block behaves exactly like the
    original template (same formulas, same look) regardless of how many
    lots come before it."""
    shift = dst_top - src_top
    for row_idx in range(src_top, src_bottom + 1):
        dst_row = row_idx + shift
        src_height = src_ws.row_dimensions[row_idx].height
        if src_height is not None:
            dst_ws.row_dimensions[dst_row].height = src_height
        for col_idx in range(min_col, max_col + 1):
            col = get_column_letter(col_idx)
            src = src_ws[f"{col}{row_idx}"]
            if isinstance(src, MergedCell):
                continue
            dst = dst_ws[f"{col}{dst_row}"]
            if isinstance(src.value, str) and src.value.startswith("="):
                # NOT Translator: each lot's block is meant to be fully
                # independent, so even $-anchored refs (e.g. "$J$9") must
                # move with the block - Translator deliberately leaves those
                # in place (that's correct for an ordinary Excel copy-paste,
                # wrong here, since it would leave every lot silently
                # pointing at lot 1's header cells). insertion_row=0 shifts
                # every row reference unconditionally, dollar or not.
                dst.value = _shift_formula_rows(src.value, insertion_row=0, shift=shift)
            else:
                dst.value = src.value
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)

    existing = {str(rng) for rng in dst_ws.merged_cells.ranges}
    for merged_range in list(src_ws.merged_cells.ranges):
        if merged_range.min_row < src_top or merged_range.max_row > src_bottom:
            continue
        new_range = (f"{get_column_letter(merged_range.min_col)}{merged_range.min_row + shift}:"
                     f"{get_column_letter(merged_range.max_col)}{merged_range.max_row + shift}")
        if new_range not in existing:
            dst_ws.merge_cells(new_range)


def _fill_block_kz(ws, block_first: int, header: dict, items: list[dict]) -> int:
    """Заполняет один блок лота (шапка+позиции+итоги) в template_kz-kz.xlsx,
    начиная со строки block_first (в оригинальном шаблоне блок начинается со
    строки 2). Возвращает номер последней строки, реально занятой блоком -
    нужно, чтобы посчитать, с какой строки начинать следующий лот."""
    off = block_first - 2

    def r(n):
        return n + off

    ws[f"B{r(2)}"] = header["manager"]
    ws[f"B{r(3)}"] = header["supplier"]
    ws[f"R{r(2)}"] = f'лот {header["lot_number"]}'
    ws[f"B{r(4)}"] = f'Дата:{header["calc_date"]}'
    ws[f"B{r(5)}"] = f'Дата начала лота:{header["lot_start"]}'
    ws[f"B{r(6)}"] = f'Дата окончания лота: {header["lot_end"]}'
    ws[f"B{r(7)}"] = f'Срок производства: {header["lead_time_days"]} дн'
    ws[f"J{r(9)}"] = header["markup_coef"]
    ws[f"K{r(9)}"] = header["usd_rate"]
    ws[f"H{r(18)}"] = header.get("road_cost") or 0

    first_row = r(12)
    extra = len(items) - 1
    total_row = r(13)
    sum_cols = ["D", "F", "G", "H", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]

    if extra > 0:
        ws.insert_rows(first_row + 1, amount=extra)
        _shift_all_formulas_below(ws, insertion_row=first_row + 1, shift=extra,
                                   max_row=block_first + 2000, max_col=22)
        for rr in range(first_row + 1, first_row + 1 + extra):
            _copy_row_style(ws, first_row, rr, min_col=2, max_col=20)  # B..T
        total_row += extra

    for i, item in enumerate(items):
        rr = first_row + i
        ws[f"B{rr}"] = i + 1
        ws[f"C{rr}"] = compose_item_name(item)
        ws[f"D{rr}"] = item["qty"]
        ws[f"E{rr}"] = item["purchase_price_ddp"]
        ws[f"T{rr}"] = item.get("extra_cost") or 0

        # J is normally the formula "=E{rr}*$J$9" (purchase price × lot
        # markup coefficient), copied down from the template's row 12 by
        # _copy_row_style above for every row past the first. When the
        # customer names their own price for this specific item, that
        # formula is exactly backwards - the price is the given fact and
        # the coefficient is what you'd back into, not the other way round.
        # Overwriting J with the plain value (not a formula) makes every
        # downstream column that reads J{rr} (K, L, M, N, O, P, Q, R, S)
        # correctly recompute off the customer's price instead, while
        # leaving $J$9 and every other item's formula untouched. Italic
        # flags it in the file as "typed in", not "computed", for whoever
        # reviews the Excel later.
        sale_price_manual = item.get("sale_price_manual") or 0
        if sale_price_manual > 0:
            j_cell = ws[f"J{rr}"]
            j_cell.value = sale_price_manual
            old_font = j_cell.font
            j_cell.font = Font(name=old_font.name, size=old_font.size, bold=old_font.bold,
                                italic=True, color=old_font.color)

    last_item_row = first_row + len(items) - 1
    for col in sum_cols:
        ws[f"{col}{total_row}"] = f"=SUM({col}{first_row}:{col}{last_item_row})"

    return r(24) + extra


def _fill_block_foreign(ws, block_first: int, header: dict, items: list[dict]) -> int:
    """Аналог _fill_block_kz() для template_foreign.xlsx (в оригинальном
    шаблоне блок начинается со строки 2, заканчивается строкой 26)."""
    off = block_first - 2

    def r(n):
        return n + off

    ws[f"B{r(2)}"] = header["manager"]
    ws[f"B{r(3)}"] = header["supplier"]
    ws[f"AE{r(2)}"] = f'лот {header["lot_number"]}'
    ws[f"B{r(4)}"] = f'Дата:{header["calc_date"]}'
    ws[f"B{r(5)}"] = f'Дата начала лота:{header["lot_start"]}'
    ws[f"B{r(6)}"] = f'Дата окончания лота: {header["lot_end"]}'
    ws[f"B{r(7)}"] = f'Срок производства: {header["lead_time_days"]} дн'
    ws[f"B{r(8)}"] = f'Срок поставки: {header.get("delivery_days", 30)} календарных дней'

    ws[f"L{r(9)}"] = header["markup_coef"]
    ws[f"O{r(9)}"] = header.get("vat_rate", 0.16)
    ws[f"H{r(9)}"] = header.get("road_cost") or 0
    ws[f"AD{r(8)}"] = header.get("currency", "USD")
    ws[f"AF{r(9)}"] = header["usd_rate"]

    first_row = r(12)
    extra = len(items) - 1
    total_row = r(13)
    sum_cols = ["E", "G", "I", "K", "M", "N", "O", "P", "Q", "S", "U", "V", "W", "Y",
                "AB", "AC", "AD", "AF", "AI"]

    if extra > 0:
        ws.insert_rows(first_row + 1, amount=extra)
        _shift_all_formulas_below(ws, insertion_row=first_row + 1, shift=extra,
                                   max_row=block_first + 2000, max_col=36)
        for rr in range(first_row + 1, first_row + 1 + extra):
            _copy_row_style(ws, first_row, rr, min_col=2, max_col=35)  # B..AI
        total_row += extra

    for i, item in enumerate(items):
        rr = first_row + i
        ws[f"B{rr}"] = i + 1
        ws[f"C{rr}"] = compose_item_name(item)
        ws[f"D{rr}"] = item.get("unit") or ""
        ws[f"E{rr}"] = item["qty"]
        ws[f"F{rr}"] = item["price_fca"]
        ws[f"P{rr}"] = item.get("overhead") or 0
        ws[f"AA{rr}"] = item["sale_price_kzt"]
        ws[f"AG{rr}"] = item["duty_rate"]
        ws[f"AH{rr}"] = item["truck_count"]
        ws[f"AI{rr}"] = item.get("extra_cost") or 0

    last_item_row = first_row + len(items) - 1
    for col in sum_cols:
        ws[f"{col}{total_row}"] = f"=SUM({col}{first_row}:{col}{last_item_row})"

    ws[f"B{r(19)}"] = f"Страна происхождения:{_compose_lot_note(items, 'country')}"
    ws[f"B{r(20)}"] = f"ТНВЭД: {_compose_lot_note(items, 'tnved')}"
    ws[f"B{r(21)}"] = f"Кол-во подвижных / вид транспорта: {_compose_lot_note(items, 'transport')}"
    if len(items) > 1:
        for coord in (f"B{r(19)}", f"B{r(20)}", f"B{r(21)}"):
            ws[coord].alignment = Alignment(wrap_text=True, vertical="top")

    # The template's "Block 1" column headers (R11..Y11) and the Q23 label
    # hardcode "$" (and, in one spot the template already had inconsistent -
    # W11 says "eur" instead of "$") regardless of which currency the lot
    # actually uses. Since every lot can now have its own currency, these
    # need to say the *right* one instead of always implying USD/EUR.
    currency = header.get("currency", "USD")
    for col in ("R", "S", "T", "U", "X", "Y"):
        cell = ws[f"{col}{r(11)}"]
        if isinstance(cell.value, str):
            cell.value = cell.value.replace("$", currency)
    w11 = ws[f"W{r(11)}"]
    if isinstance(w11.value, str):
        w11.value = w11.value.replace("eur", currency)
    ws[f"Q{r(23)}"] = currency

    return r(26) + extra


def fill_lots_kz(template_path: str, output_path: str, lots: list[dict]):
    """
    template_kz-kz.xlsx, один или несколько лотов одного тендера в одном
    Excel-файле. lots - список {"header": {...}, "items": [...]}, где header
    и items - те же поля, что раньше принимала fill_multi(). Каждый лот -
    полная копия шаблона (шапка+позиции+итоги), одна под другой в одном
    листе, с отступом между лотами.
    """
    wb = openpyxl.load_workbook(template_path)  # без data_only! формулы должны остаться
    ws = wb["расчет"]
    ref_ws = openpyxl.load_workbook(template_path)["расчет"]  # пристинная копия для дублирования блока

    block_first, block_last, gap = 2, 24, 2
    cursor = block_first
    for lot in lots:
        _copy_block(ws, ref_ws, block_first, block_last, cursor, min_col=1, max_col=20)
        block_actual_last = _fill_block_kz(ws, cursor, lot["header"], lot["items"])
        cursor = block_actual_last + 1 + gap

    wb.save(output_path)
    print(f"Saved: {output_path} (lots: {len(lots)})")


def fill_lots_foreign(template_path: str, output_path: str, lots: list[dict]):
    """
    template_foreign.xlsx, один или несколько лотов одного тендера в одном
    Excel-файле. lots - список {"header": {...}, "items": [...]}, те же поля,
    что раньше принимала fill_multi_foreign(). Каждый лот - полная копия
    шаблона, одна под другой в одном листе.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb["расчет"]
    ref_ws = openpyxl.load_workbook(template_path)["расчет"]

    block_first, block_last, gap = 2, 26, 2
    cursor = block_first
    for lot in lots:
        _copy_block(ws, ref_ws, block_first, block_last, cursor, min_col=2, max_col=35)
        block_actual_last = _fill_block_foreign(ws, cursor, lot["header"], lot["items"])
        cursor = block_actual_last + 1 + gap

    wb.save(output_path)
    print(f"Saved: {output_path} (lots: {len(lots)})")


if __name__ == "__main__":
    from datetime import date

    example_lots = [
        {
            "header": {
                "manager": "Батыр", "supplier": "RG GOLD", "lot_number": "T-0003701",
                "calc_date": date.today().strftime("%d.%m.%Y"),
                "lot_start": "19.06.2026", "lot_end": "25.06.2026",
                "lead_time_days": 10, "markup_coef": 1.5, "usd_rate": 486.19,
            },
            "items": [
                {"name": "Песок (отсев) + доставка", "qty": 150, "purchase_price_ddp": 8500},
            ],
        },
        {
            "header": {
                "manager": "Батыр", "supplier": "ТОО Другой Поставщик", "lot_number": "T-0003702",
                "calc_date": date.today().strftime("%d.%m.%Y"),
                "lot_start": "26.06.2026", "lot_end": "02.07.2026",
                "lead_time_days": 15, "markup_coef": 1.3, "usd_rate": 486.19,
            },
            "items": [
                {"name": "Щебень фракции 20-40", "qty": 300, "purchase_price_ddp": 6200},
            ],
        },
    ]
    fill_lots_kz("template_kz-kz.xlsx", "output_filled.xlsx", example_lots)
