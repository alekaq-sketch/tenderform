"""
Heat Energy - Streamlit tender calculator.
"""

import os
import tempfile
import traceback
from datetime import date
from io import BytesIO

import openpyxl
import pandas as pd
import streamlit as st
from streamlit.errors import StreamlitSecretNotFoundError

from extract_heuristic import extract_items as extract_items_free
from extract_raw import extract_any
from extract_with_llm import extract_fields as extract_items_llm
from fetch_usd_rate import get_usd_rate
from fill_tender_template import fill_multi


st.set_page_config(
    page_title="Heat Energy · Tender Calculator",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600&family=IBM+Plex+Mono:wght@400;500;600&family=Inter:wght@400;500;600&display=swap');

    :root {
        --bg: #0a0e17;
        --paper: #f7f3e9;
        --paper-soft: #f2ebdb;
        --paper-line: #d8cea9;
        --ink: #1e252d;
        --ink-soft: #58606a;
        --accent: #186b5d;
        --accent-soft: #3e8d7b;
        --steel: #2d6480;
        --highlight: #f9d28d;
        --danger: #b94e2b;
        --font-display: "Fraunces", Georgia, serif;
        --font-mono: "IBM Plex Mono", "Courier New", monospace;
        --font-body: "Inter", -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }

    .stApp {
        background:
            radial-gradient(circle at 20% 45%, rgba(24, 107, 93, 0.1) 0%, transparent 40%),
            radial-gradient(circle at 80% 80%, rgba(45, 100, 128, 0.09) 0%, transparent 42%),
            linear-gradient(135deg, #0a0e17 0%, #1a2332 52%, #0f1a28 100%);
        color: var(--paper);
        font-family: var(--font-body);
    }

    [data-testid="stHeader"], [data-testid="stToolbar"] {
        background: transparent;
    }

    [data-testid="stSidebar"], [data-testid="collapsedControl"] {
        display: none;
    }

    .block-container {
        max-width: 1060px;
        padding: 24px 28px 72px;
    }

    .topbar {
        display: flex;
        justify-content: space-between;
        align-items: flex-start;
        gap: 24px;
        margin: 0 0 28px;
    }

    .brand {
        font-family: var(--font-display);
        font-size: 29px;
        line-height: 1.02;
        letter-spacing: 0;
    }

    .brand span {
        display: block;
    }

    .brand .accent {
        color: var(--highlight);
        font-weight: 600;
    }

    .brand-subtitle {
        margin-top: 8px;
        color: rgba(255, 255, 255, 0.82);
        font-size: 13px;
    }

    .stamp {
        width: 94px;
        height: 94px;
        border: 2px dashed var(--highlight);
        border-radius: 50%;
        color: var(--highlight);
        transform: rotate(-7deg);
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        font-family: var(--font-mono);
        flex: 0 0 auto;
    }

    .stamp .label {
        color: var(--accent-soft);
        font-size: 8px;
        text-transform: uppercase;
        letter-spacing: 0.12em;
    }

    .stamp .lot {
        max-width: 78px;
        overflow-wrap: anywhere;
        text-align: center;
        font-size: 13px;
        font-weight: 600;
    }

    .stamp .dt {
        color: var(--accent-soft);
        font-size: 9px;
        margin-top: 2px;
    }

    .sheet {
        background: linear-gradient(180deg, var(--paper) 0%, var(--paper-soft) 100%);
        color: var(--ink);
        border-radius: 8px;
        box-shadow: 0 30px 60px -20px rgba(0, 0, 0, 0.55);
        padding: 8px;
    }

    .step {
        padding: 30px 34px;
        border-bottom: 1px solid var(--paper-line);
    }

    .step-final {
        border-bottom: 0;
        text-align: center;
        padding: 34px 34px 38px;
    }

    .step-title {
        display: flex;
        align-items: baseline;
        gap: 12px;
        margin: 0 0 6px;
        font-family: var(--font-display);
        font-size: 23px;
        font-weight: 600;
        color: #07152a;
    }

    .step-no {
        font-family: var(--font-mono);
        font-size: 12px;
        color: var(--accent);
        letter-spacing: 0.08em;
    }

    .step-hint {
        margin: 0 0 18px;
        color: var(--ink-soft);
        font-size: 13.5px;
    }

    .stFileUploader section {
        background: rgba(255, 255, 255, 0.36);
        border: 1.5px dashed var(--paper-line);
        border-radius: 6px;
        padding: 22px 18px;
    }

    .stFileUploader label, .stTextInput label, .stNumberInput label {
        color: var(--ink-soft) !important;
        font-size: 12.5px !important;
    }

    .stTextInput input, .stNumberInput input {
        background: #fff;
        border: 1px solid var(--paper-line);
        border-radius: 4px;
        color: var(--ink);
    }

    .stTextInput input:focus, .stNumberInput input:focus {
        border-color: var(--steel);
        box-shadow: 0 0 0 1px var(--steel);
    }

    .stButton button, .stDownloadButton button {
        border-radius: 4px;
        font-family: var(--font-body);
        font-weight: 600;
    }

    .stButton button[kind="primary"], .stDownloadButton button[kind="primary"] {
        background: linear-gradient(135deg, var(--accent) 0%, var(--accent-soft) 100%);
        border: none;
        color: white;
        min-width: 268px;
        min-height: 50px;
        font-family: var(--font-display);
        font-size: 17px;
    }

    .stButton button[kind="secondary"] {
        border: 1px solid var(--steel);
        color: var(--steel);
        background: transparent;
    }

    .stDataFrame, [data-testid="stDataEditor"] {
        border: 1px solid var(--paper-line);
        border-radius: 6px;
        overflow: hidden;
    }

    .status-ok {
        color: var(--accent);
        font-size: 13.5px;
        margin: 10px 0 0;
    }

    .helper-card {
        margin-top: 14px;
        padding: 14px 16px;
        border: 1px solid var(--paper-line);
        border-radius: 6px;
        background: rgba(255, 255, 255, 0.58);
        color: var(--ink-soft);
        font-size: 12.5px;
    }

    div[data-testid="stAlert"] {
        border-radius: 6px;
    }

    hr {
        border-color: var(--paper-line);
    }

    @media (max-width: 700px) {
        .block-container {
            padding: 18px 14px 48px;
        }

        .topbar {
            flex-direction: column;
        }

        .stamp {
            align-self: flex-end;
        }

        .sheet {
            padding: 4px;
        }

        .step {
            padding: 24px 18px;
        }
    }
</style>
""",
    unsafe_allow_html=True,
)


def init_state() -> None:
    if "items" not in st.session_state:
        st.session_state["items"] = []
    if "header" not in st.session_state:
        st.session_state["header"] = {
            "manager": "",
            "supplier": "",
            "lot_number": "",
            "calc_date": date.today().strftime("%d.%m.%Y"),
            "lot_start": "",
            "lot_end": "",
            "lead_time_days": 10,
            "markup_coef": 1.5,
            "usd_rate": 0.0,
            "road_cost": 0.0,
        }
    if "raw_text" not in st.session_state:
        st.session_state["raw_text"] = ""
    if "template_file" not in st.session_state:
        st.session_state["template_file"] = None
    if "generated_file" not in st.session_state:
        st.session_state["generated_file"] = None
    if "generated_name" not in st.session_state:
        st.session_state["generated_name"] = "расчёт.xlsx"


def format_date(value: str) -> str:
    digits = "".join(filter(str.isdigit, value))[:8]
    if len(digits) <= 2:
        return digits
    if len(digits) <= 4:
        return f"{digits[:2]}.{digits[2:]}"
    return f"{digits[:2]}.{digits[2:4]}.{digits[4:]}"


@st.cache_data
def fetch_usd_cached() -> float | None:
    try:
        return get_usd_rate()
    except Exception as exc:
        st.warning(f"Не удалось получить курс: {exc}")
        return None


def normalize_item(item: dict) -> dict:
    return {
        "name": item.get("name") or "",
        "unit": item.get("unit") or "",
        "qty": item.get("qty") or 0,
        "unit_price": item.get("unit_price") or 0,
    }


def get_secret(name: str, default: str = "") -> str:
    try:
        return st.secrets.get(name, default)
    except StreamlitSecretNotFoundError:
        return default


def save_uploaded_file(uploaded_file) -> str:
    suffix = os.path.splitext(uploaded_file.name)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getbuffer())
        return tmp.name


def build_excel_from_template(template_path: str, header: dict, items: list[dict]) -> bytes:
    payload_items = [
        {
            "item_name": item["name"],
            "qty": item["qty"],
            "purchase_price_ddp": item["unit_price"],
        }
        for item in items
        if item.get("name") and str(item["name"]).strip()
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        out_path = tmp.name

    try:
        fill_multi(template_path, out_path, header, payload_items)
        with open(out_path, "rb") as file:
            return file.read()
    finally:
        if os.path.exists(out_path):
            os.unlink(out_path)


def build_excel_from_uploaded_template(template_file, header: dict, items: list[dict]) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(template_file.getbuffer()))
    ws = wb["расчет"] if "расчет" in wb.sheetnames else wb[wb.sheetnames[0]]

    ws["B2"] = header["manager"]
    ws["B3"] = header["supplier"]
    ws["R2"] = f'лот {header["lot_number"]}'
    ws["B4"] = f'Дата:{header["calc_date"]}'
    ws["B5"] = f'Дата начала лота:{header["lot_start"]}'
    ws["B6"] = f'Дата окончания лота: {header["lot_end"]}'
    ws["B7"] = f'Срок производства: {header["lead_time_days"]} дн'
    ws["J9"] = header["markup_coef"]
    ws["K9"] = header["usd_rate"]
    ws["H18"] = header["road_cost"]

    first_row = 12
    for index, item in enumerate(items):
        if item.get("name"):
            row = first_row + index
            ws[f"B{row}"] = index + 1
            ws[f"C{row}"] = item["name"]
            ws[f"D{row}"] = item["qty"]
            ws[f"E{row}"] = item["unit_price"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def has_valid_items(items: list[dict]) -> bool:
    return bool(items) and any(item.get("name") for item in items)


init_state()
items = st.session_state["items"]
header = st.session_state["header"]

stamp_lot = header["lot_number"] or "T-0000001"
stamp_date = header["calc_date"] or date.today().strftime("%d.%m.%Y")

st.markdown(
    f"""
<div class="topbar">
    <div>
        <div class="brand">
            <span>Heat Energy</span>
            <span class="accent">Tender Calculator</span>
        </div>
        <div class="brand-subtitle">Локальный калькулятор расчётов и подготовки Excel</div>
    </div>
    <div class="stamp">
        <div class="label">лот №</div>
        <div class="lot">{stamp_lot}</div>
        <div class="dt">{stamp_date}</div>
    </div>
</div>
<div class="sheet">
""",
    unsafe_allow_html=True,
)

st.markdown(
    """
<div class="step">
    <h2 class="step-title"><span class="step-no">01</span>Загрузите документ</h2>
    <p class="step-hint">ТЗ или техническая спецификация — .docx, .xlsx или .pdf.</p>
""",
    unsafe_allow_html=True,
)

with st.container():
    uploaded_file = st.file_uploader(
        "Техническое задание или спецификация",
        type=["docx", "xlsx", "pdf"],
        label_visibility="collapsed",
    )

    if uploaded_file:
        st.markdown(f'<p class="status-ok">Файл выбран: {uploaded_file.name}</p>', unsafe_allow_html=True)
        if st.button("Распознать позиции", use_container_width=False):
            tmp_path = save_uploaded_file(uploaded_file)
            try:
                with st.spinner("Читаю документ..."):
                    raw_text = extract_any(tmp_path)
                    extracted_items = extract_items_free(tmp_path)
                st.session_state["raw_text"] = raw_text[:8000]
                st.session_state["items"] = [normalize_item(item) for item in extracted_items]
                st.session_state["generated_file"] = None
                st.success(f"Найдено позиций: {len(st.session_state['items'])}. Проверьте таблицу ниже.")
                st.rerun()
            except Exception as exc:
                st.error(f"Ошибка распознавания: {exc}")
                traceback.print_exc()
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

    if st.session_state["raw_text"]:
        with st.expander("Сырой текст документа для проверки"):
            st.text_area("", st.session_state["raw_text"], height=220, label_visibility="collapsed")

    llm_col, llm_btn_col = st.columns([0.78, 0.22])
    with llm_col:
        api_key = st.text_input(
            "Anthropic API ключ",
            value=get_secret("ANTHROPIC_API_KEY"),
            type="password",
            placeholder="Anthropic API ключ (опционально, для сложных документов)",
            label_visibility="collapsed",
        )
    with llm_btn_col:
        run_llm = st.button("Распознать через LLM", use_container_width=True)

    if run_llm:
        if not api_key:
            st.warning("Введите Anthropic API ключ.")
        elif not st.session_state["raw_text"]:
            st.warning("Сначала загрузите и распознайте документ.")
        else:
            try:
                with st.spinner("Claude анализирует документ..."):
                    result = extract_items_llm(st.session_state["raw_text"], api_key=api_key)
                st.session_state["items"] = [normalize_item(item) for item in result.get("items", [])]
                st.session_state["generated_file"] = None
                st.success(f"LLM нашёл позиций: {len(st.session_state['items'])}.")
                st.rerun()
            except Exception as exc:
                st.error(f"Ошибка LLM: {exc}")

    st.markdown('<div class="helper-card">Шаблон Excel для офлайн-режима</div>', unsafe_allow_html=True)
    template_file = st.file_uploader(
        "Шаблон Excel для офлайн-режима",
        type=["xlsx"],
        key="template_upload",
        label_visibility="collapsed",
    )
    if template_file:
        st.session_state["template_file"] = template_file
        st.caption("Если основной шаблон недоступен, можно использовать загруженный файл.")

st.markdown("</div>", unsafe_allow_html=True)

st.markdown(
    """
<div class="step">
    <h2 class="step-title"><span class="step-no">02</span>Проверьте позиции</h2>
    <p class="step-hint">Найденные значения можно править прямо в таблице.</p>
""",
    unsafe_allow_html=True,
)

current_items = st.session_state["items"]
if not current_items:
    current_items = [{"name": "", "unit": "", "qty": 0, "unit_price": 0}]

items_df = pd.DataFrame(current_items, columns=["name", "unit", "qty", "unit_price"])
edited_df = st.data_editor(
    items_df,
    use_container_width=True,
    hide_index=True,
    num_rows="dynamic",
    column_config={
        "name": st.column_config.TextColumn("Наименование", width="large"),
        "unit": st.column_config.TextColumn("Ед. изм.", width="small"),
        "qty": st.column_config.NumberColumn("Кол-во", step=1, format="%g"),
        "unit_price": st.column_config.NumberColumn("Цена DDP", step=0.01, format="%.2f"),
    },
    key="items_editor",
)

st.session_state["items"] = [
    normalize_item(row)
    for row in edited_df.fillna("").to_dict("records")
    if any(str(value).strip() for value in row.values())
]
items = st.session_state["items"]

st.markdown("</div>", unsafe_allow_html=True)

st.markdown(
    """
<div class="step">
    <h2 class="step-title"><span class="step-no">03</span>Данные лота</h2>
""",
    unsafe_allow_html=True,
)

left_col, right_col = st.columns(2)
with left_col:
    header["manager"] = st.text_input("Менеджер", value=header["manager"], placeholder="Алишер")
    header["lot_number"] = st.text_input("Номер лота", value=header["lot_number"], placeholder="T-0000001")
    calc_date = st.text_input("Дата расчёта", value=header["calc_date"], placeholder="дд.мм.гггг")
    header["calc_date"] = format_date(calc_date)
    lot_end = st.text_input("Окончание лота", value=header["lot_end"], placeholder="дд.мм.гггг")
    header["lot_end"] = format_date(lot_end)
    header["road_cost"] = st.number_input("Сумма дорожных расходов", value=float(header["road_cost"]), step=1000.0)

with right_col:
    header["supplier"] = st.text_input("Поставщик", value=header["supplier"], placeholder="RG GOLD")
    header["markup_coef"] = st.number_input("Коэффициент наценки", value=float(header["markup_coef"]), step=0.05)
    lot_start = st.text_input("Начало лота", value=header["lot_start"], placeholder="дд.мм.гггг")
    header["lot_start"] = format_date(lot_start)
    header["lead_time_days"] = st.number_input(
        "Срок производства, дней",
        value=int(header["lead_time_days"]),
        step=1,
    )

usd_col, usd_btn_col = st.columns([0.84, 0.16])
with usd_col:
    header["usd_rate"] = st.number_input("Курс USD", value=float(header["usd_rate"]), step=0.01)
with usd_btn_col:
    st.write("")
    st.write("")
    if st.button("с Нацбанка РК", use_container_width=True):
        rate = fetch_usd_cached()
        if rate:
            header["usd_rate"] = rate
            st.success(f"Курс USD: {rate}")
            st.rerun()

st.markdown("</div>", unsafe_allow_html=True)

st.markdown('<div class="step-final">', unsafe_allow_html=True)
generate_col, download_col = st.columns([1, 1])

with generate_col:
    if st.button("Сформировать расчёт", type="primary", use_container_width=True):
        if not has_valid_items(items):
            st.error("Добавьте хотя бы одну позицию.")
        elif header["usd_rate"] <= 0:
            st.error("Укажите курс USD.")
        elif not os.path.exists("template.xlsx"):
            st.error("Шаблон template.xlsx не найден в репозитории.")
        else:
            try:
                st.session_state["generated_file"] = build_excel_from_template("template.xlsx", header, items)
                st.session_state["generated_name"] = f"расчёт_{header['lot_number'] or 'лот'}.xlsx"
                st.success("Готово. Формулы пересчитаются при открытии в Excel.")
            except Exception as exc:
                st.error(f"Ошибка: {exc}")
                traceback.print_exc()

with download_col:
    template_file = st.session_state["template_file"]
    if st.button("Сформировать из загруженного шаблона", use_container_width=True):
        if not template_file:
            st.warning("Выберите шаблон Excel в первом блоке.")
        elif not has_valid_items(items):
            st.error("Добавьте хотя бы одну позицию.")
        elif header["usd_rate"] <= 0:
            st.error("Укажите курс USD.")
        else:
            try:
                st.session_state["generated_file"] = build_excel_from_uploaded_template(template_file, header, items)
                st.session_state["generated_name"] = f"расчёт_{header['lot_number'] or 'лот'}.xlsx"
                st.success("Готово. Файл можно скачать.")
            except Exception as exc:
                st.error(f"Ошибка: {exc}")

if st.session_state["generated_file"]:
    st.download_button(
        "Скачать Excel",
        data=st.session_state["generated_file"],
        file_name=st.session_state["generated_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

st.markdown("</div></div>", unsafe_allow_html=True)
