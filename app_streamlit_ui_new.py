"""Heat Energy - Streamlit tender calculator (UI rewrite).

Goal: remove brittle HTML/CSS wrappers (topbar/step/layout blocks) and rebuild the
UI with Streamlit-native components so that widget colors/sizes remain consistent.
"""

import os
import tempfile
import traceback
from datetime import date
from io import BytesIO

import openpyxl
import pandas as pd
import streamlit as st
from streamlit.errors import StreamlitSecretNotFoundError

from extract_heuristic import extract_items as extract_items_free
from extract_raw import extract_any
from extract_with_llm import extract_fields as extract_items_llm
from fetch_usd_rate import get_usd_rate
from fill_tender_template import fill_multi


st.set_page_config(
    page_title="Heat Energy · Tender Calculator",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Minimal safe CSS: tweak border-radius only.
st.markdown(
    """
<style>
  div[data-testid="stTextInput"] input,
  div[data-testid="stNumberInput"] input,
  div[data-testid="stTextArea"] textarea {
    border-radius: 6px;
  }

  section[data-testid="stDataEditor"],
  div[data-testid="stDataEditor"] {
    border-radius: 10px;
    border: 1px solid rgba(0,0,0,0.08);
  }
</style>
""",
    unsafe_allow_html=True,
)


def init_state() -> None:
    st.session_state.setdefault("items", [])
    st.session_state.setdefault(
        "header",
        {
            "manager": "",
            "supplier": "",
            "lot_number": "",
            "calc_date": date.today().strftime("%d.%m.%Y"),
            "lot_start": "",
            "lot_end": "",
            "lead_time_days": 10,
            "markup_coef": 1.5,
            "usd_rate": 0.0,
            "road_cost": 0.0,
        },
    )
    st.session_state.setdefault("raw_text", "")
    st.session_state.setdefault("template_file", None)
    st.session_state.setdefault("generated_file", None)
    st.session_state.setdefault("generated_name", "расчёт.xlsx")


def format_date(value: str) -> str:
    digits = "".join(filter(str.isdigit, value))[:8]
    if len(digits) <= 2:
        return digits
    if len(digits) <= 4:
        return f"{digits[:2]}.{digits[2:]}"
    return f"{digits[:2]}.{digits[2:4]}.{digits[4:]}"


@st.cache_data
def fetch_usd_cached() -> float | None:
    try:
        return get_usd_rate()
    except Exception as exc:
        st.warning(f"Не удалось получить курс: {exc}")
        return None


def normalize_item(item: dict) -> dict:
    return {
        "name": item.get("name") or "",
        "unit": item.get("unit") or "",
        "qty": item.get("qty") or 0,
        "unit_price": item.get("unit_price") or 0,
    }


def get_secret(name: str, default: str = "") -> str:
    try:
        return st.secrets.get(name, default)
    except StreamlitSecretNotFoundError:
        return default


def save_uploaded_file(uploaded_file) -> str:
    suffix = os.path.splitext(uploaded_file.name)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getbuffer())
        return tmp.name


def build_excel_from_template(template_path: str, header: dict, items: list[dict]) -> bytes:
    payload_items = [
        {
            "item_name": item["name"],
            "qty": item["qty"],
            "purchase_price_ddp": item["unit_price"],
        }
        for item in items
        if item.get("name") and str(item["name"]).strip()
    ]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        out_path = tmp.name

    try:
        fill_multi(template_path, out_path, header, payload_items)
        with open(out_path, "rb") as file:
            return file.read()
    finally:
        if os.path.exists(out_path):
            os.unlink(out_path)


def build_excel_from_uploaded_template(template_file, header: dict, items: list[dict]) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(template_file.getbuffer()))
    ws = wb["расчет"] if "расчет" in wb.sheetnames else wb[wb.sheetnames[0]]

    ws["B2"] = header["manager"]
    ws["B3"] = header["supplier"]
    ws["R2"] = f'лот {header["lot_number"]}'
    ws["B4"] = f'Dата:{header["calc_date"]}'
    ws["B5"] = f'Dата начала лота:{header["lot_start"]}'
    ws["B6"] = f'Dата окончания лота: {header["lot_end"]}'
    ws["B7"] = f'Срок производства: {header["lead_time_days"]} дн'
    ws["J9"] = header["markup_coef"]
    ws["K9"] = header["usd_rate"]
    ws["H18"] = header["road_cost"]

    first_row = 12
    for index, item in enumerate(items):
        if item.get("name"):
            row = first_row + index
            ws[f"B{row}"] = index + 1
            ws[f"C{row}"] = item["name"]
            ws[f"D{row}"] = item["qty"]
            ws[f"E{row}"] = item["unit_price"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def has_valid_items(items: list[dict]) -> bool:
    return bool(items) and any(item.get("name") for item in items)


init_state()
items = st.session_state["items"]
header = st.session_state["header"]

# Header
left, right = st.columns([3, 1])
with left:
    st.title("Heat Energy — Tender Calculator")
    st.caption("Локальный калькулятор расчётов и подготовки Excel")
with right:
    st.metric("Лот", header.get("lot_number") or "T-0000001")
    st.metric("Дата", header.get("calc_date") or date.today().strftime("%d.%m.%Y"))

st.divider()

# Step 1
st.subheader("01. Загрузите документ")
step1_left, step1_right = st.columns([0.7, 0.3])

with step1_left:
    uploaded_file = st.file_uploader(
        "ТЗ или спецификация",
        type=["docx", "xlsx", "pdf"],
        accept_multiple_files=False,
    )

    if uploaded_file:
        st.success(f"Файл выбран: {uploaded_file.name}")

    if st.button("Распознать", use_container_width=True, disabled=uploaded_file is None):
        tmp_path = save_uploaded_file(uploaded_file)
        try:
            with st.spinner("Читаю документ..."):
                raw_text = extract_any(tmp_path)
                extracted_items = extract_items_free(tmp_path)
            st.session_state["raw_text"] = raw_text[:8000]
            st.session_state["items"] = [normalize_item(item) for item in extracted_items]
            st.session_state["generated_file"] = None
            st.success(f"Найдено позиций: {len(st.session_state['items'])}.")
            st.rerun()
        except Exception as exc:
            st.error(f"Ошибка распознавания: {exc}")
            traceback.print_exc()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    if st.session_state["raw_text"]:
        with st.expander("Сырой текст документа (для проверки)", expanded=False):
            st.text_area("", st.session_state["raw_text"], height=220, label_visibility="collapsed")

    template_file = st.file_uploader(
        "Шаблон Excel (xlsx)",
        type=["xlsx"],
        key="template_upload",
    )
    if template_file is not None:
        st.session_state["template_file"] = template_file

with step1_right:
    st.subheader("LLM")
    api_key = st.text_input(
        "Anthropic API ключ",
        value=get_secret("ANTHROPIC_API_KEY"),
        type="password",
        placeholder="опционально",
    )

    if st.button("Распознать через LLM", use_container_width=True):
        if not api_key:
            st.warning("Введите Anthropic API ключ.")
        elif not st.session_state["raw_text"]:
            st.warning("Сначала загрузите и распознайте документ.")
        else:
            try:
                with st.spinner("Claude анализирует документ..."):
                    result = extract_items_llm(st.session_state["raw_text"], api_key=api_key)
                st.session_state["items"] = [normalize_item(item) for item in result.get("items", [])]
                st.session_state["generated_file"] = None
                st.success(f"LLM нашёл позиций: {len(st.session_state['items'])}.")
                st.rerun()
            except Exception as exc:
                st.error(f"Ошибка LLM: {exc}")

# Step 2
st.divider()
st.subheader("02. Проверьте позиции")

current_items = st.session_state["items"] or [{"name": "", "unit": "", "qty": 0, "unit_price": 0}]
items_df = pd.DataFrame(current_items, columns=["name", "unit", "qty", "unit_price"])

edited_df = st.data_editor(
    items_df,
    use_container_width=True,
    hide_index=True,
    num_rows="dynamic",
    column_config={
        "name": st.column_config.TextColumn("Наименование", width="large"),
        "unit": st.column_config.TextColumn("Ед. изм.", width="small"),
        "qty": st.column_config.NumberColumn("Кол-во", step=1, format="%g"),
        "unit_price": st.column_config.NumberColumn("Цена DDP", step=0.01, format="%.2f"),
    },
    key="items_editor",
)

st.session_state["items"] = [
    normalize_item(row)
    for row in edited_df.fillna("").to_dict("records")
    if any(str(value).strip() for value in row.values())
]
items = st.session_state["items"]

# Step 3
st.divider()
st.subheader("03. Данные лота")

left_col, right_col = st.columns(2)
with left_col:
    header["manager"] = st.text_input("Менеджер", value=header["manager"], placeholder="Алишер")
    header["lot_number"] = st.text_input("Номер лота", value=header["lot_number"], placeholder="T-0000001")

    calc_date_raw = st.text_input("Дата расчёта", value=header["calc_date"], placeholder="дд.мм.гггг")
    header["calc_date"] = format_date(calc_date_raw)

    road_cost = st.number_input(
        "Сумма дорожных расходов",
        value=float(header["road_cost"]),
        step=1000.0,
    )
    header["road_cost"] = road_cost

with right_col:
    header["supplier"] = st.text_input("Поставщик", value=header["supplier"], placeholder="RG GOLD")
    header["markup_coef"] = st.number_input("Коэффициент наценки", value=float(header["markup_coef"]), step=0.05)

    lot_start_raw = st.text_input("Начало лота", value=header["lot_start"], placeholder="дд.мм.гггг")
    header["lot_start"] = format_date(lot_start_raw)

    lot_end_raw = st.text_input("Окончание лота", value=header["lot_end"], placeholder="дд.мм.гггг")
    header["lot_end"] = format_date(lot_end_raw)

    header["lead_time_days"] = st.number_input("Срок производства, дней", value=int(header["lead_time_days"]), step=1)

usd_col, usd_btn_col = st.columns([0.85, 0.15])
with usd_col:
    header["usd_rate"] = st.number_input("Курс USD", value=float(header["usd_rate"]), step=0.01)
with usd_btn_col:
    if st.button("с Нацбанка РК", use_container_width=True):
        rate = fetch_usd_cached()
        if rate:
            header["usd_rate"] = rate
            st.success(f"Курс USD: {rate}")
            st.rerun()

# Generate
st.divider()

gen_col, dl_col = st.columns(2)
with gen_col:
    if st.button("Сформировать расчёт", type="primary", use_container_width=True):
        if not has_valid_items(items):
            st.error("Добавьте хотя бы одну позицию.")
        elif header.get("usd_rate", 0) <= 0:
            st.error("Укажите курс USD.")
        elif not os.path.exists("template.xlsx"):
            st.error("Шаблон template.xlsx не найден в репозитории.")
        else:
            try:
                st.session_state["generated_file"] = build_excel_from_template("template.xlsx", header, items)
                st.session_state["generated_name"] = f"расчёт_{header['lot_number'] or 'лот'}.xlsx"
                st.success("Готово. Формулы пересчитаются при открытии в Excel.")
            except Exception as exc:
                st.error(f"Ошибка: {exc}")
                traceback.print_exc()

with dl_col:
    if st.button("Сформировать из загруженного шаблона", use_container_width=True):
        template_file = st.session_state.get("template_file")
        if not template_file:
            st.warning("Выберите шаблон Excel в первом блоке.")
        elif not has_valid_items(items):
            st.error("Добавьте хотя бы одну позицию.")
        elif header.get("usd_rate", 0) <= 0:
            st.error("Укажите курс USD.")
        else:
            try:
                st.session_state["generated_file"] = build_excel_from_uploaded_template(template_file, header, items)
                st.session_state["generated_name"] = f"расчёт_{header['lot_number'] or 'лот'}.xlsx"
                st.success("Готово. Файл можно скачать.")
            except Exception as exc:
                st.error(f"Ошибка: {exc}")

if st.session_state.get("generated_file"):
    st.download_button(
        "Скачать Excel",
        data=st.session_state["generated_file"],
        file_name=st.session_state["generated_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

